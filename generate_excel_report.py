#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script để tạo file Excel tổng hợp kết quả LAB 4
"""

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Tạo workbook
wb = Workbook()
wb.remove(wb.active)  # Xóa sheet mặc định

# ============ SHEET 1: BÀI 1 - ĐIỂM =============
ws1 = wb.create_sheet('Bài 1 - Điểm')

# Tiêu đề
ws1['A1'] = 'BÀI 1: PHÂN TÍCH ĐIỂM HỌC PHẦN'
ws1['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws1['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws1.merge_cells('A1:E1')

# Dữ liệu
scores = np.array([
    [8.0, 7.5, 8.5, 7.0],
    [6.5, 6.0, 7.0, 6.5],
    [9.0, 8.5, 9.0, 8.5],
    [5.0, 5.5, 6.0, 5.5],
    [7.5, 7.0, 8.0, 7.5],
    [4.5, 5.0, 5.5, 5.0],
    [8.5, 9.0, 8.0, 9.0],
    [6.0, 6.5, 6.0, 6.5],
    [7.0, 7.5, 7.0, 8.0],
    [9.5, 9.0, 9.5, 9.0]
])
weights = np.array([0.1, 0.2, 0.3, 0.4])
final_score = scores @ weights

def classify_grade(score):
    if score >= 8.5:
        return 'A'
    elif score >= 7.0:
        return 'B'
    elif score >= 5.5:
        return 'C'
    elif score >= 4.0:
        return 'D'
    else:
        return 'F'

grades = np.array([classify_grade(score) for score in final_score])

# Header
headers1 = ['SV', 'Chuyên cần', 'Giữa kỳ', 'Thực hành', 'Cuối kỳ', 'Tổng kết', 'Loại']
for col, header in enumerate(headers1, 1):
    cell = ws1.cell(row=3, column=col)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Dữ liệu
for i in range(len(scores)):
    ws1.cell(row=4+i, column=1).value = f'SV{i+1}'
    for j in range(4):
        ws1.cell(row=4+i, column=2+j).value = scores[i][j]
    ws1.cell(row=4+i, column=6).value = final_score[i]
    ws1.cell(row=4+i, column=7).value = grades[i]

# Độ rộng cột
for col in range(1, 8):
    ws1.column_dimensions[get_column_letter(col)].width = 12

# ============ SHEET 2: BÀI 2 - CHUYÊN CẦN =============
ws2 = wb.create_sheet('Bài 2 - Chuyên cần')

ws2['A1'] = 'BÀI 2: PHÂN TÍCH CHUYÊN CẦN'
ws2['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws2['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws2.merge_cells('A1:I1')

attendance = np.array([
    [1,1,1,1,1,1,1,1],
    [1,1,0,1,1,0,1,1],
    [1,0,0,1,1,1,0,1],
    [1,1,1,1,0,1,1,1],
    [0,1,1,0,1,1,1,0],
    [1,1,1,1,1,1,0,1],
    [1,0,1,0,1,0,1,0],
    [1,1,1,1,1,1,1,0],
    [0,0,1,1,0,1,1,1],
    [1,1,1,0,1,1,1,1],
    [1,1,0,0,1,0,1,1],
    [1,1,1,1,1,0,1,1]
])

present_count = attendance.sum(axis=1)
rate = present_count / attendance.shape[1] * 100

headers2 = ['SV'] + [f'Buổi {i+1}' for i in range(8)] + ['Tổng', 'Tỉ lệ %']
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

for i in range(len(attendance)):
    ws2.cell(row=4+i, column=1).value = f'SV{i+1}'
    for j in range(8):
        ws2.cell(row=4+i, column=2+j).value = attendance[i][j]
    ws2.cell(row=4+i, column=10).value = present_count[i]
    ws2.cell(row=4+i, column=11).value = rate[i]

for col in range(1, 12):
    ws2.column_dimensions[get_column_letter(col)].width = 10

# ============ SHEET 3: BÀI 3 - DOANH THU =============
ws3 = wb.create_sheet('Bài 3 - Doanh thu')

ws3['A1'] = 'BÀI 3: PHÂN TÍCH DOANH THU'
ws3['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws3['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws3.merge_cells('A1:F1')

sales = np.array([
    [120, 150, 130, 140, 160],
    [125, 145, 128, 142, 158],
    [130, 155, 135, 150, 162],
    [135, 160, 140, 152, 168],
    [140, 165, 145, 155, 170],
    [138, 162, 142, 153, 169],
    [145, 170, 150, 160, 175]
])

daily_total = sales.sum(axis=1)
product_total = sales.sum(axis=0)
product_mean = sales.mean(axis=0)
product_std = sales.std(axis=0)

headers3 = ['Ngày'] + [f'SP{i+1}' for i in range(5)] + ['Tổng ngày']
for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=col)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

for i in range(len(sales)):
    ws3.cell(row=4+i, column=1).value = f'Ngày {i+1}'
    for j in range(5):
        ws3.cell(row=4+i, column=2+j).value = sales[i][j]
    ws3.cell(row=4+i, column=7).value = daily_total[i]

# Tổng sản phẩm
ws3.cell(row=12, column=1).value = 'TỔNG'
ws3.cell(row=12, column=1).font = Font(bold=True)
for j in range(5):
    ws3.cell(row=12, column=2+j).value = product_total[j]
    ws3.cell(row=12, column=2+j).font = Font(bold=True)

ws3.cell(row=13, column=1).value = 'TRUNG BÌNH'
ws3.cell(row=13, column=1).font = Font(bold=True)
for j in range(5):
    ws3.cell(row=13, column=2+j).value = round(product_mean[j], 2)
    ws3.cell(row=13, column=2+j).font = Font(bold=True)

ws3.cell(row=14, column=1).value = 'ĐỘ LỆCH CHUẨN'
ws3.cell(row=14, column=1).font = Font(bold=True)
for j in range(5):
    ws3.cell(row=14, column=2+j).value = round(product_std[j], 2)
    ws3.cell(row=14, column=2+j).font = Font(bold=True)

for col in range(1, 8):
    ws3.column_dimensions[get_column_letter(col)].width = 12

# ============ SHEET 4: BÀI 4 - TỒN KHO =============
ws4 = wb.create_sheet('Bài 4 - Tồn kho')

ws4['A1'] = 'BÀI 4: QUẢN LÝ TỒN KHO'
ws4['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws4['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws4.merge_cells('A1:F1')

stock = np.array([35, 8, 12, 5, 40, 18, 7, 22, 9, 15])
min_stock = np.array([20, 15, 15, 10, 25, 20, 12, 18, 12, 15])
price = np.array([30, 25, 28, 22, 35, 20, 18, 24, 19, 21])
need_import = np.maximum(min_stock - stock, 0)
cost = need_import * price
status = np.where(stock < min_stock, 'Thiếu', 'Đủ')

headers4 = ['Hàng', 'Tồn hiện tại', 'Tồn tối thiểu', 'Cần nhập', 'Giá nhập', 'Chi phí', 'Trạng thái']
for col, header in enumerate(headers4, 1):
    cell = ws4.cell(row=3, column=col)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

for i in range(len(stock)):
    ws4.cell(row=4+i, column=1).value = f'HH{i+1}'
    ws4.cell(row=4+i, column=2).value = stock[i]
    ws4.cell(row=4+i, column=3).value = min_stock[i]
    ws4.cell(row=4+i, column=4).value = need_import[i]
    ws4.cell(row=4+i, column=5).value = price[i]
    ws4.cell(row=4+i, column=6).value = cost[i]
    ws4.cell(row=4+i, column=7).value = status[i]
    
    # Highlight thiếu hàng
    if need_import[i] > 0:
        ws4.cell(row=4+i, column=7).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# Tổng
ws4.cell(row=14, column=1).value = 'TỔNG CHI PHÍ'
ws4.cell(row=14, column=1).font = Font(bold=True)
ws4.cell(row=14, column=6).value = cost.sum()
ws4.cell(row=14, column=6).font = Font(bold=True)

for col in range(1, 8):
    ws4.column_dimensions[get_column_letter(col)].width = 14

# ============ SHEET 5: SO SÁNH =============
ws5 = wb.create_sheet('So sánh')

ws5['A1'] = 'SO SÁNH TỔNG HỢPHAI MIỀN'
ws5['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws5['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws5.merge_cells('A1:C1')

comparison_data = [
    ['Tiêu chí', 'Bài 1 - Điểm', 'Bài 3 - Doanh thu'],
    ['Cấu trúc', '10×4', '7×5'],
    ['Axis 0 (hàng)', 'Sinh viên', 'Ngày'],
    ['Axis 1 (cột)', '4 thành phần', '5 sản phẩm'],
    ['Phép aggregation', 'Matrix mult', 'Sum axis'],
    ['Boolean indexing', 'NHIỀU', 'ÍT'],
    ['Tính phân hóa', 'Cao (std, z-score)', 'Vừa phải (std)'],
    ['Mục tiêu', 'Quản lý/Cảnh báo', 'Xu hướng/Quy hoạch'],
]

for row_idx, row_data in enumerate(comparison_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws5.cell(row=row_idx, column=col_idx)
        cell.value = value
        if row_idx == 3:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

for col in range(1, 4):
    ws5.column_dimensions[get_column_letter(col)].width = 20

# ============ LƯU FILE =============
wb.save('H:\\HocTap\\Chuyende3\\lab4_doanthingocanh\\LAB4_TONG_HOP_KET_QUA.xlsx')
print("✓ Đã tạo file Excel: LAB4_TONG_HOP_KET_QUA.xlsx")
print("✓ File được lưu trong thư mục: H:\\HocTap\\Chuyende3\\lab4_doanthingocanh\\")

